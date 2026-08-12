from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from scanner.models import ScanSummary


def export_csv(summary: ScanSummary, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Adresse IP", "Nom d'hôte", "Port", "Service", "Produit", "Version", "CPE", "CVE", "CVSS max", "KEV", "Confiance", "Risque", "Description", "Recommandation"])
        for finding in summary.findings:
            cve_text = ", ".join(cve.cve_id for cve in finding.cves)
            if not cve_text and not finding.fingerprint.version:
                cve_text = "CVE non évaluables : version du produit non identifiée"
            writer.writerow([
                finding.ip,
                finding.hostname,
                finding.port,
                finding.service,
                finding.fingerprint.product,
                finding.fingerprint.version,
                finding.fingerprint.cpe,
                cve_text,
                max((cve.cvss_score for cve in finding.cves), default=0) or "",
                "Oui" if any(cve.kev for cve in finding.cves) else "Non",
                finding.fingerprint.confidence,
                finding.risk,
                finding.description,
                finding.recommendation,
            ])
    return path


def export_excel(summary: ScanSummary, path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Résultats"
    rows = [
        ["Cible", summary.target],
        ["Durée (s)", round(summary.duration_seconds, 2)],
        ["Appareils détectés", summary.host_count],
        ["Ports ouverts", summary.open_port_count],
        ["CVE détectées", summary.cve_count],
        ["CVE critiques", summary.critical_cve_count],
        ["CVE CISA KEV", summary.kev_count],
        [],
        [
            "Adresse IP",
            "Nom d'hôte",
            "Port",
            "Service",
            "Produit",
            "Version",
            "CPE",
            "CVE",
            "CVSS max",
            "KEV",
            "Confiance",
            "Risque port",
            "Description",
            "Recommandation",
        ],
    ]
    for row in rows:
        ws.append(row)
    for finding in summary.findings:
        max_cvss = max((cve.cvss_score for cve in finding.cves), default=0)
        ws.append([
            finding.ip,
            finding.hostname,
            finding.port,
            finding.service,
            finding.fingerprint.product,
            finding.fingerprint.version,
            finding.fingerprint.cpe,
            ", ".join(cve.cve_id for cve in finding.cves),
            max_cvss,
            "Oui" if any(cve.kev for cve in finding.cves) else "Non",
            finding.fingerprint.confidence,
            finding.risk,
            finding.description,
            finding.recommendation,
        ])
    for cell in ws[9]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="12263A")
    for column in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 70)
        ws.column_dimensions[column[0].column_letter].width = width
    cve_ws = wb.create_sheet("CVE")
    cve_ws.append(["CVE", "Score", "Sévérité", "KEV", "Confiance", "Produit", "Version", "CWE", "Publication", "Modification", "Description", "Références"])
    for finding in summary.findings:
        for cve in finding.cves:
            cve_ws.append([
                cve.cve_id,
                cve.cvss_score,
                cve.severity,
                "Oui" if cve.kev else "Non",
                cve.confidence,
                finding.fingerprint.product,
                finding.fingerprint.version,
                cve.cwe,
                cve.published,
                cve.last_modified,
                cve.description,
                "\n".join(cve.references),
            ])
    for cell in cve_ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="12263A")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def export_pdf(summary: ScanSummary, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(str(path), pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = [
        Paragraph("NetScope Scanner - Rapport d'analyse réseau", styles["Title"]),
        Spacer(1, 10),
        Paragraph(
            f"Cible : {summary.target} | Durée : {summary.duration_seconds:.2f}s | "
            f"Appareils : {summary.host_count} | Ports ouverts : {summary.open_port_count} | "
            f"CVE : {summary.cve_count} | CVE critiques : {summary.critical_cve_count} | KEV : {summary.kev_count}",
            styles["Normal"],
        ),
        Spacer(1, 12),
        Paragraph(
            "Avertissement : une CVE est proposée uniquement lorsqu'un produit ou une version a été identifié. "
            "Ces résultats restent potentiels et peuvent contenir des faux positifs.",
            styles["Italic"],
        ),
        Spacer(1, 12),
    ]
    data = [["IP", "Port", "Service", "Produit", "Version", "CVE", "CVSS", "KEV", "Confiance"]]
    for finding in summary.findings:
        max_cvss = max((cve.cvss_score for cve in finding.cves), default=0)
        data.append([
            finding.ip,
            str(finding.port),
            finding.service,
            finding.fingerprint.product or "-",
            finding.fingerprint.version or "-",
            ", ".join(cve.cve_id for cve in finding.cves) or "-",
            str(max_cvss or "-"),
            "Oui" if any(cve.kev for cve in finding.cves) else "Non",
            finding.fingerprint.confidence,
        ])
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#12263A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D8E1EF")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 14))
    story.append(Paragraph("Détails CVE", styles["Heading2"]))
    for finding in summary.findings:
        for cve in finding.cves:
            story.append(Paragraph(f"{cve.cve_id} - {cve.severity} - CVSS {cve.cvss_score}", styles["Heading3"]))
            story.append(Paragraph(f"Produit : {finding.fingerprint.product} {finding.fingerprint.version} | Confiance : {cve.confidence} | KEV : {'Oui' if cve.kev else 'Non'}", styles["Normal"]))
            story.append(Paragraph(cve.description[:1200], styles["Normal"]))
            story.append(Spacer(1, 8))
    doc.build(story)
    return path
